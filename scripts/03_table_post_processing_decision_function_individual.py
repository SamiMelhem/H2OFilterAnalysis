
"""
This script combines information from the following output categories:

- Data that was filtered out of the raw input excel files for defined scenarios (N < 3, N < 20, all C < detection limit (DL), etc)
    (this file generated in script 01_data_filtering.py)
- The remaining data that passed each filter from the original raw input that was then passed through trend analysis and summary stats
    (file generated in script 02_trend_analysis_filtered_data.py)
- Information coupling constituents to their groupings (as defined in permit documents): VOCs, Primary MCLs, Notification Levels, CECs, etc...

Resulting data can be used to assess and propose an Action based on stats and counts.
Can also be used to construct Categories as deliverables to client

authors: rweatherl, smelhem

"""

# import glob
# import itertools
# import matplotlib.pyplot as plt
from os import getcwd
from os.path import dirname, join
from numpy import where#, int64
from pandas import concat, read_csv, to_numeric, DataFrame#, MultiIndex
from matplotlib import use
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# from xlwings import Book

use('Qt5Agg')

permit_name =  'MFSG'  ## 'ARC WDR'  ##'MFSG'

## User input data. Change every time scripts 01_ and 02_ are re-run!
filterdata_input = 'MFSG_allfilters_summary_ALL_rw_20230609.csv'
statsdata_input = 'MFSG_summary_stats_ALL_rw_20230609.csv'

## User input/output directories
## cwd = getcwd()+'\\Amendment 5'  ## this does not work for Robin. We just want to be in scripts directory.
cwd = getcwd()
input_dir = join(dirname(cwd), 'generated_input')
data_dir = join(dirname(cwd), 'output', f'{permit_name}')
output_dir = join(dirname(cwd), 'output', f'{permit_name}')


### bring in Group and legal limit data for sorting. class_constituent_ordering.csv is a living doc and subject to changes
sorting = read_csv(join(input_dir, 'class_constituent_ordering.csv'),
                      index_col='DB Name', encoding= 'unicode_escape')
sorting.index = sorting.index.str.strip()           ## remove hidden spaces
sorting.index = sorting.index.rename('Constituent_sort')
sorting['Name in Permit'] = sorting['Name in Permit'].str.strip()       ## remove hidden spaces
sorting.drop_duplicates(subset = 'Name in Permit', keep = 'first', inplace=True)
#sorting['Regulatory Limit']=to_numeric(sorting['Regulatory Limit'], errors='coerce')


def read_filterdata():

    ##import data
    fd_raw = read_csv(join(data_dir, filterdata_input), index_col = 'C_wo_units')

    fd = fd_raw.copy()
    fd['Site'] = fd['Site'].astype('int64')
    fd.index = fd.index.rename('Constituent_sort')
    # fd.rename({'C_wo_units': 'Constituent_sort'}, inplace=True)
    fdsort = fd.join(sorting, on='Constituent_sort', how='outer', rsuffix='_legal')  ## how='inner'

    fdsort.dropna(subset='Site', inplace=True)

    return fd, fdsort

def read_statsdata():

    sd_raw = read_csv(join(data_dir, statsdata_input), index_col = 'c_nounits')

    sd = sd_raw.copy()

    sd.index = sd.index.rename('Constituent_sort')
    sd['Site'] = sd['Site'].astype('int64')
    sdsort = sd.join(sorting, on='Constituent_sort', how='outer', rsuffix='_legal')  #how='inner'
    sdsort.drop(['mann-kendall stats'], axis=1, inplace=True)
    # sdsort['Category'].fillna('Other', inplace=True)

    sdsort.dropna(subset='Site', inplace=True)

    return sd, sdsort

def merge_data(fdsort, sdsort):  ##fd, sd

    #all data merged

    combined = concat([sdsort, fdsort])
    # combined.dropna(subset='Site', inplace = True)   ## NEEDS TO BE DEBUGGED!!!

    combined.sort_values(by=['Category', 'Constituent', 'Site'], inplace=True)
    # df.set_index(['Constituent_sort', 'Site'], inplace = True)
    combined['Reg Limit'] = to_numeric(combined['Reg Limit'], errors='coerce')

    ## add snippet here to convert pg/L, ug/L and ng/L to mg/L
    # combined['Converted Units'] = combined['DB Units']
    combined['Converted Units'] = combined['Units']  ## use units directly from raw database, not from class_constituent_ordering.csv

    pg = combined['Units'] == 'pg/L'
    combined.loc[pg, 'Min':'DL max'] = combined.loc[pg, 'Min':'DL max']*1E-9
    # combined.loc[pg, 'DL min':'DL max'] = combined.loc[pg, 'DL min':'DL max'] * 1E-9
    combined.loc[pg, 'Converted Units'] = 'mg/L'

    ng = combined['Units'] == 'ng/L'
    combined.loc[ng, 'Min':'DL max'] = combined.loc[ng, 'Min':'DL max'] * 1E-6
    # combined.loc[ng, 'DL min': 'DL max'] = combined.loc[ng, 'DL min':'DL max'] * 1E-6
    combined.loc[ng, 'Converted Units'] = 'mg/L'

    ug = combined['Units'] == 'ug/L'
    combined.loc[ug, 'Min':'DL max'] = combined.loc[ug, 'Min':'DL max'] * 1E-3
    combined.loc[ug, 'DL min':'DL max'] = combined.loc[ug, 'DL min':'DL max'] * 1E-3
    combined.loc[ug, 'Converted Units'] = 'mg/L'

    combined['DL min'] = combined['DL min'].apply(lambda x: '{:.6f}'.format(x))
    combined['DL max'] = combined['DL max'].apply(lambda x: '{:.6f}'.format(x))
    combined['DL'] = where(combined['DL min'] == combined['DL max'], combined['DL max'],
                              combined['DL min'].astype(str) + '-' + combined['DL max'].astype(str))

    combined.set_index(['Constituent_sort', 'Site'], inplace=True)

    df = combined.copy()[['Units', 'Converted Units', 'Category', 'Reg Limit', 'Reg Unit', 'Date Range', 'Count', 'DL', 'Methods',
                          'Nr. Non-Detects', 'Detection Rate', 'Min', '25th %ile', 'Mean', 'Median', '75th %ile',
                          'Max', 'DateOfMax', 'trend', 's', 'p', 'Action', 'Notes', 'Notes_legal']]

    # df.to_csv(path.join(data_dir, f'{permit_name}_sorted_testing.csv')) ##sanity check

    return df


def decision_function(df):

    ## when detection frequency is low
    lowfreq_na = (df['Reg Limit'].isna()) & (df['Notes'] == 'Pending manual review') ## no reg. limit
    df.loc[lowfreq_na, 'Notes'] = 'Low detection rate + no reg. limit'
    df.loc[lowfreq_na, 'Action'] = 'Reduce Frequency'

    lowfreq_notna = (df['Reg Limit'].notna()) & (df['Notes'] == 'Pending manual review') & (df['Max'] < df['Reg Limit']) ## max < reg. limit
    df.loc[lowfreq_notna, 'Notes'] = 'Low detection rate + max < reg. limit'
    df.loc[lowfreq_notna, 'Action'] = 'Reduce Frequency'

    lowfreq_rl = (df['Reg Limit'].notna()) & (df['Notes'] == 'Pending manual review') & (df['Max'] > df['Reg Limit'])  ## max > reg. limit
    df.loc[lowfreq_rl, 'Notes'] = 'Low detection rate + max > reg. limit'
    df.loc[lowfreq_rl, 'Action'] = 'Keep Same Frequency'

    # add condition for when <4 samples are above detection limit (type 3 data)
    # when <4 samples are detected:
    #   reduce frequency when max val is < regulatory limit, or when there is no regulatory limit
    #   keep same frequency when max val is >= regulatory limit
    # lessthan4_below = (df['Reg Limit'].notna()) &  ((df['Count'] - df['Nr. Non-Detects']) <4) & (df['Max'] < df['Reg Limit'])
    # lessthan4_above = (df['Reg Limit'].notna()) &  ((df['Count'] - df['Nr. Non-Detects']) <4) & (df['Max'] >= df['Reg Limit'])
    # lessthan4_norl = (df['Reg Limit'].isna()) &  ((df['Count'] - df['Nr. Non-Detects']) <4)

    lessthan4_below = (df['Reg Limit'].notna()) & (df['Notes'] == 'Less than 4 values > DL') & (df['Max'] < df['Reg Limit'])
    lessthan4_above = (df['Reg Limit'].notna()) & (df['Notes'] == 'Less than 4 values > DL') & (df['Max'] >= df['Reg Limit'])
    lessthan4_norl = (df['Reg Limit'].isna()) & (df['Notes'] == 'Less than 4 values > DL')
    df.loc[lessthan4_below, 'Notes'] = '<4 samples above detection limit, max val < reg. limit'
    df.loc[lessthan4_below, 'Action'] = 'Reduce Frequency'
    df.loc[lessthan4_above, 'Notes'] = '<4 samples above detection limit, max val >= reg. limit'
    df.loc[lessthan4_above, 'Action'] = 'Keep Same Frequency'
    df.loc[lessthan4_norl, 'Notes'] = '<4 samples above detection limit, no reg. limit'
    df.loc[lessthan4_norl, 'Action'] = 'Reduce Frequency'

    
    ## NO REGULATORY LIMIT EXISTS

    ## statistical analysis, no reguatory limit
    stats_na_inc = (df['Reg Limit'].isna()) & (df['trend'] == 'increasing')      ## increasing trend
    df.loc[stats_na_inc, 'Notes'] = 'Increasing trend, no reg. limit'
    df.loc[stats_na_inc, 'Action'] = 'Keep Same Frequency'

    stats_na_notrend = (df['Reg Limit'].isna()) & (df['trend'] == 'no trend')  ## no trend
    df.loc[stats_na_notrend, 'Notes'] = 'No trend, no reg. limit'
    df.loc[stats_na_notrend, 'Action'] = 'Reduce Frequency'

    stats_na_dec = (df['Reg Limit'].isna()) & (df['trend'] == 'decreasing')   ## decreasing trend
    df.loc[stats_na_dec, 'Notes'] = 'Decreasing trend, no reg. limit'
    df.loc[stats_na_dec, 'Action'] = 'Reduce Frequency'

    ## REGULATORY LIMIT EXISTS

    ## max > 50% reg limit, not outlier (DateofMax < 5 yrs ago -- 01/01/2018)
    ## ---> keep current frequency
    maxrecent_rl = (df['Reg Limit'].notna()) & (df['Max'] > (0.5*df['Reg Limit'])) & (df['DateOfMax'] > '2018-01-01')  #this is correct logic, decision #1 for Type 4 with reg limits -- Rach
    df.loc[maxrecent_rl, 'Notes'] = 'Max > 50% reg limit, within last 5 yrs'
    df.loc[maxrecent_rl, 'Action'] = 'Keep Same Frequency'


    ## max > 50% reg limit, outlier (DateofMax > 5 yrs ago -- 01/01/2018)   #these are correct logic, decisions #2 for Type 4 with reg limits -- Rach
    high_max_inc =  (df['Reg Limit'].notna()) & (df['Max'] > (0.5*df['Reg Limit'])) & (df['DateOfMax'] < '2018-01-01') & (df['trend'] == 'increasing')
    high_max_notrend = (df['Reg Limit'].notna()) & (df['Max'] > (0.5*df['Reg Limit'])) & (df['DateOfMax'] < '2018-01-01') & (df['trend'] == 'no trend')
    high_max_dec = (df['Reg Limit'].notna()) & (df['Max'] > (0.5*df['Reg Limit'])) & (df['DateOfMax'] < '2018-01-01') & (df['trend'] == 'decreasing')
    df.loc[high_max_inc, 'Notes'] = 'Max > 50% reg. limit, outlier > 5 yr ago, increasing trend'
    df.loc[high_max_inc, 'Action'] = 'Keep Same Frequency'
    df.loc[high_max_notrend, 'Notes'] = 'Max > 50% reg. limit, outlier > 5 yr ago, no trend'
    df.loc[high_max_notrend, 'Action'] = 'Reduce Frequency'
    df.loc[high_max_dec, 'Notes'] = 'Max > 50% reg. limit, outlier > 5 yr ago, decreasing trend'
    df.loc[high_max_dec, 'Action'] = 'Reduce Frequency'

    ## max < 50% reg limit, no date necessary just look at RL and Trend  #i believe that this is correct for decision #3 for Type 4 with reg limits -- Rach
    low_max_inc = (df['Reg Limit'].notna()) &  (df['Max'] < (0.5*df['Reg Limit'])) & (df['trend'] == 'increasing')
    low_max_notrend = (df['Reg Limit'].notna()) &  (df['Max'] < (0.5*df['Reg Limit']))  & (df['trend'] == 'no trend')
    low_max_dec = (df['Reg Limit'].notna()) &  (df['Max'] < (0.5*df['Reg Limit'])) & (df['trend'] == 'decreasing')
    df.loc[low_max_inc, 'Notes'] = 'Max < 50% reg. limit, increasing trend'
    df.loc[low_max_inc, 'Action'] = 'Keep Same Frequency'
    df.loc[low_max_notrend, 'Notes'] = 'Max < 50% reg. limit, no trend'
    df.loc[low_max_notrend, 'Action'] = 'Reduce Frequency'
    df.loc[low_max_dec, 'Notes'] = 'Max < 50% reg. limit, decreasing trend'
    df.loc[low_max_dec, 'Action'] = 'Reduce Frequency'


    ## "Type 0" constituents in groundwater -- reccomend elimination
    groundwater = False
    if groundwater:
        type0 = df['Notes_legal'].str.contains('Type 0', na=False)
        df.loc[type0, 'Notes'] = 'Eliminate'

    return df

def read_to_implement(wb, orig_ws, values, row, new_ws_num):
    if orig_ws.cell(row, 2).value in values[new_ws_num-1][1:] or new_ws_num == 13 or row == 1: # [0] to get the first title's list of rows, and [1:] to ignore the row, column increments
            for c in range(1, orig_ws.max_column+1):
                wb.worksheets[new_ws_num].cell(values[new_ws_num-1][0], c).value = orig_ws.cell(row, c).value
            values[new_ws_num-1][0] += 1
            return True
    return False

def clear_empty_rows(ws):
    '''
    Clears any rows that are completely empty throughout the worksheet given
    Input: ws: Worksheet
    '''
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            ws.delete_rows(row[0].row)

def insert_in_order(wb, titles, ws_num):
    list_of_constituents = list(titles.values())[ws_num-1][1:] # creates the list of constituents by grabbing the right list and excluding the first number
    # Collects the placements of each constituent within the alphabetized constituents
    placements = [next((i for i in range(1, wb.worksheets[ws_num].max_row+1) if wb.worksheets[ws_num].cell(i,2).value == constituent), None) for constituent in list_of_constituents] 
    # Inserts the needed amount of rows for the newly ordered constituents
    wb.worksheets[ws_num].insert_rows(2, len(list_of_constituents))
    for row in range(2, len(list_of_constituents)+2): # Loops through the placement of each ordered constituent
        for col in range(1, wb.worksheets[ws_num].max_column+1): # Loops through all the values within that constituent row
            if placements[row-2] != None:
                wb.worksheets[ws_num].cell(row, col).value = wb.worksheets[ws_num].cell(placements[row-2]+len(list_of_constituents), col).value
                wb.worksheets[ws_num].cell(placements[row-2]+len(list_of_constituents), col).value = None
            else:
                wb.worksheets[ws_num].cell(row, 2).value = list_of_constituents[row-2]
                break
    clear_empty_rows(wb.worksheets[ws_num])

def clear_unnecessary(wb, ws_num):
    for row in range(2, wb.worksheets[ws_num].max_row + 1):
        # clears brackets and quotes
        if wb.worksheets[ws_num].cell(row, 12).value != None:
            wb.worksheets[ws_num].cell(row, 12).value = str(wb.worksheets[ws_num].cell(row, 12).value)[2:-2]

        # rounds up or down detection limits
        if wb.worksheets[ws_num].cell(row, 14).value != None: 
            wb.worksheets[ws_num].cell(row, 14).value = round(float(wb.worksheets[ws_num].cell(row, 14).value))

        # rounds the p-values to the thousandths place
        if wb.worksheets[ws_num].cell(row, 24).value != None and len(str(wb.worksheets[ws_num].cell(row, 24).value)) > 1:
            wb.worksheets[ws_num].cell(row, 24).value = round(float(wb.worksheets[ws_num].cell(row, 24).value), 3) 

if __name__ == "__main__":

    fd, fdsort = read_filterdata()
    sd, sdsort = read_statsdata()
    df = merge_data(fdsort,sdsort)
    df['Methods'] = df['Methods'].str.replace("[","").str.replace("]","").str.replace("' '", ", ").str.replace("'","")

    df_decision = decision_function(df)

    ## QA. Should be empty.
    t = df_decision.loc[df_decision['Action'] == 'TO REVIEW']

    # df.to_csv(join(output_dir, f'{permit_name}_post_processed_v10.csv'))

    ## can just do this directly and remove the df.to_csv line above.
    # excel = read_csv(join(output_dir, f'{permit_name}_post_processed_v10.csv')).to_excel(join(output_dir, f'{permit_name}_post_processed_v11.xlsx'))
    ofile = f'{permit_name}_post_processed_ALL_rw_20230609.xlsx'
    df_decision.to_excel(join(output_dir, ofile))




    wb = load_workbook(join(output_dir, ofile))
    titles = { # title: [currRow, rowNums...]
        'Inorganic Chemicals w Prim MCLs' : [1, 'Aluminum', 'Antimony', 'Arsenic', 'Asbestos', 'Barium', 'Beryllium', 'Cadmium', 'Total Chromium', 'Hexavalent Chromium (Cr VI)',
                                             'Cyanide', 'Fluoride', 'Mercury', 'Nickel', 'Perchlorate', 'Selenium', 'Thallium'],
        'Constituents with SMCLs' : [1, 'Aluminum', 'Apparent Color','Copper', 'Surfactants', 'Iron', 'Manganese', 'Methyl Tert Butyl Ether (MTBE)',
                                     'Odor', 'Silver', 'Thiobencarb', 'Turbidity', 'Zinc',],
        'Radioactivity' : [1, 'Radium 226 and 228', 'Gross Alpha particle activity (excluding radon and uranium)', 'Tritium', 'Strontium 90',
                           'Gross Beta', 'Uranium'],
        'Regulated Organic Chemicals' : [1, 'Benzene', 'Carbon Tetrachloride', 'o-Dichlorobenzene (1,2-DCB)', 'p-Dichlorobenzene',
                                         '1,1-Dichloroethane', '1,2-Dichloroethane', '1,1-Dichloroethylene', 'cis-1,2-Dichloroethylene',
                                         'trans-1,2-Dichloroethylene', 'Methylene Chloride', '1,2-Dichloropropane', '1,3-Dichloropropene (Total)',
                                         'Ethyl Benzene', 'Methyl Tert Butyl Ether (MTBE)', 'Chlorobenzene', 'Styrene',
                                         '1,1,2,2-Tetrachloroethane', 'Tetrachloroethylene (PCE)', 'Toluene', '1,2,4-Trichlorobenzene',
                                         '1,1,1-Trichloroethane', '1,1,2-Trichloroethane', 'Trichloroethylene (TCE)',
                                         'Fluorotrichloromethane (Freon11)', 'Trichlorotrifluoroethane (Freon 113)',
                                         'Vinyl chloride (VC)', 'm,p-Xylenes'],
        'Non-Volatile SOCs' : [1, 'Alachlor', 'Atrazine', 'Bentazon', 'Benzo (a) Pyrene', 'Carbofuran (Furadan)', 'Chlordane',
                               '2,4-D', 'Dalapon (qualitative)', 'Dibromochloropropane (DBCP)', 'Di (2-Ethylhexyl) Adipate',
                               'Di(2-ethylhexyl)phthalate (DEHP)', 'Dinoseb', 'Diquat', 'Endothall', 'Endrin',
                               'Ethylene Dibromide (EDB)', 'Glyphosate', 'Heptachlor', 'Heptachlor Epoxide',
                               'Hexachlorobenzene', 'Hexachlorocyclopentadiene', 'Lindane (Gamma-BHC)', 'Methoxychlor',
                               'Molinate', 'Oxamyl (Vydate)', 'PCP', 'Picloram', 'Polychlorinated Biphenyls (PCBs)',
                               'Simazine', 'Thiobencarb', 'Toxaphene', '1,2,3-Trichloropropane', '2,3,7,8-TCDD (Dioxin)', '2,4,5-TP (Silvex)',],
        'Disinfection Byproducts' : [1, 'Total Trihalomethanes', 'Bromodichloromethane', 'Bromoform', 'Chloroform (Trichloromethane)',
                                     'Chlorodibromomethane', 'D/DBP Haloacetic Acids (HAA5)', 'Monochloroacetic acid',
                                     'Dichloroacetic acid', 'Trichloroacetic acid', 'Monobromoacetic acid', 'Dibromoacetic acid',
                                     'Bromate', 'Chlorite by IC'],
        'Constituents with Notifi Levels' : [1, 'Boron', 'n-Butylbenzene', 'sec-Butylbenzene', 'tert-Butylbenzene', 'Carbon Disulfide',
                                             'Chlorate', 'o-Chlorotoluene', 'p-Chlorotoluene', 'Diazinon', 'Dichlorodifluoromethane (Freon 12)',
                                             '1,4-Dioxane', 'Ethylene Glycol', 'Formaldehyde', 'HMX', 'Isopropylbenzene', 'Manganese',
                                             '4-Methyl-2-Pentanone (MIBK)', 'Naphthalene', 'N-Nitrosodiethylamine (NDEA)',
                                             'N-Nitrosodimethylamine (NDMA)', 'N-Nitrosodi-N-Propylamine (NDPA)', 'PFOA', 'PFOS', 'PFHxS',
                                             'PFBS', 'Propachlor', 'n-Propylbenzene', 'RDX', 'Tertiary Butyl Alcohol (TBA)',
                                             '1,2,4-Trimethylbenzene', '1,3,5-Trimethylbenzene', '2,4,6-Trinitrotoluene (TNT)', 'Vanadium'],
        'General Physical and Minerals' : [1, 'Calcium', 'Potassium', 'Sodium', 'Hardness (Total, as CaCO3)'],
        'Pesticides' : [1, 'Aldrin', 'Dieldrin', "p,p' DDD", "p,p' DDE", "p,p' DDD", 'Endosulfan I (alpha)', 'Endosulfan II (beta)', 'Endosulfan Sulfate',
                        'Endrin Aldehyde', 'Alpha-BHC', 'Beta-BHC', 'Delta-BHC'],
        'Acid Extractables' : [1, '2,4,6-Trichlorophenol', 'Parachlorometa cresol (P-chloro-m-cresol)', '2-Chlorophenol', '2,4-Dichlorophenol',
                               '2,4-Dimethylphenol', '2-Nitrophenol', '4-Nitrophenol', '2,4-Dinitrophenol',
                               '2-Methyl-4,6-Dinitrophenol', 'Phenol'],
        'Base Neutral Extractables' : [1, 'Acenaphthene, Total Weight', 'Benzidine', 'Hexachloroethane', 'Bis (2-Chloroethyl) Ether', '2-Chloronaphthalene',
                                       '1,3-Dichlorobenzene', '3,3-Dichlorobenzidine', '2,4-Dinitrotoluene', '2,6-Dinitrotoluene',
                                       '1,2-diphenylhydrazine', 'Fluoranthene', '4-Bromophenyl Phenyl Ether', '4-Chlorophenyl Phenyl Ether',
                                       'Bis (2-Chlorethoxyl) Methane', 'Bis (2-Chloroisopropyl) Ether', 'Hexachlorobutadiene', 'Isophorone',
                                       'Nitrobenzene', 'N-nitrosodiphenylamine', 'Butylbenzylphthalate', 'Di-n-Butylphthalate',
                                       'Di-N-Octylphthalate', 'Diethylphthalate', 'Dimethylphthalate', 'Benzo (a) Anthracene',
                                       'Benzo (b) Fluoranthene', 'Benzo (k) Fluoranthene', 'Chrysene', 'Acenaphthylene', 'Anthracene',
                                       'Benzo (g,h,i) Perylene', 'Fluorene', 'Phenanthrene', 'Dibenzo (a,h) Anthracene', 'Indeno (1,2,3,c,d) Pyrene',
                                       'Pyrene'],
        'Volatile Organics' : [1, 'Acrolein', 'Acrylonitrile', 'Chlorobenzene', 'Chloroethane', 'Methyl chloride (dichloromethane)',
                               'Bromomethane (Methyl Bromide)', '2-Chloroethyl Vinyl Ether'],
        'Extras' : [1]
    }

    for title in titles.keys():
        wb.create_sheet(title) # Create the workbook sheet

    # Take respective rows and put them into their places
    orig_ws = wb.worksheets[0]
    for row in range (1, orig_ws.max_row + 1):
        values = list(titles.values())

        implements = [read_to_implement(wb, orig_ws, values, row, i) for i in range (1, 13)]

        if True not in implements:
            read_to_implement(wb, orig_ws, values, row, 13)

    wb.save(join(output_dir, ofile))

    # Filter the rows alphabetically (exception: nonvolatile_socs)
    for title in titles.keys():
        ws = wb[title]
        # Extract the data from the worksheet into a DataFrame
        data = ws.values
        columns = next(data, None)
        df = DataFrame(data, columns=columns)

        # Sort the DataFrame by 'Constituent_sort'
        sorted_df = None
        if df.columns.size > 0:
            sorted_df = df.sort_values(by=df.columns[1])

            # Rewrite the sorted data back to the worksheet
            ws.delete_rows(1, ws.max_row)
            for row in dataframe_to_rows(sorted_df, index=False, header=True):
                ws.append(row)

    [insert_in_order(wb, titles, i) for i in range(1, 13)]

    [clear_empty_rows(wb[title]) for title in titles.keys()]

    # Put the needed changes here

    ## Universal changes
    [clear_unnecessary(wb, i) for i in range (1, 13)]

    ## Inorganic Chemical w Prim MCLs

    # Changes 'Total Chromium' to 'Chromium (Total)
    for i in range(2, wb.worksheets[1].max_row + 1):
        if wb.worksheets[1].cell(i, 2).value == 'Total Chromium':
            wb.worksheets[1].cell(i, 2).value = 'Chromium (Total)'
            break


    ## Constituents with SMCLs

    # Changes 'Apparent Color' to 'Color'
    for i in range(2, wb.worksheets[2].max_row + 1):
        if wb.worksheets[2].cell(i, 2).value == 'Apparent Color':
            wb.worksheets[2].cell(i, 2).value = 'Color'
            break


    ## Radioactivity

    # Changes the names from 'Radium 226 and 228' and 'Gross Alpha particle activity (excluding radon and uranium)', respectively
    both = 0
    for i in range(2, wb.worksheets[3].max_row + 1):
        if wb.worksheets[3].cell(i, 2).value == 'Radium 226 and 228':
            wb.worksheets[3].cell(i, 2).value = 'Combined Radium-226 and Radium-228'
            both += 1
        elif wb.worksheets[3].cell(i, 2).value == 'Gross Alpha particle activity (excluding radon and uranium)':
            wb.worksheets[3].cell(i, 2).value = 'Gross Alpha Particle Activity (Including Radium-226 but Excluding Radon and Uranium)'
            both += 1
        if both == 2:
            break


    ## Regulated Organic Chemicals

    # Changes the names respectively
    both = 0
    for i in range(2, wb.worksheets[4].max_row + 1):
        if wb.worksheets[4].cell(i, 2).value == 'o-Dichlorobenzene (1,2-DCB)':
            wb.worksheets[4].cell(i, 2).value = '1,2-Dichlorobenzene (1,2-DCB)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == 'p-Dichlorobenzene':
            wb.worksheets[4].cell(i, 2).value = '1,4-Dichlorobenzene (1,4-DCB)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == '1,1-Dichloroethane':
            wb.worksheets[4].cell(i, 2).value = '1,1-Dichloroethane (1,1-DCA)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == '1,2-Dichloroethane':
            wb.worksheets[4].cell(i, 2).value = '1,2-Dichloroethane (1,2-DCA)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == '1,1-Dichloroethylene':
            wb.worksheets[4].cell(i, 2).value = '1,1-Dichloroethylene (1,1-DCE)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == 'Methylene Chloride':
            wb.worksheets[4].cell(i, 2).value = 'Dichloromethane (Methylene Chloride)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == '1,3-Dichloropropene (Total)':
            wb.worksheets[4].cell(i, 2).value = '1,3-Dichloropropene (1,3-DCP)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == 'Chlorobenzene':
            wb.worksheets[4].cell(i, 2).value = 'Monochlorobenzene (Chlorobenzene)'
            both += 1
        elif wb.worksheets[4].cell(i, 2).value == '1,1,1-Trichloroethane':
            wb.worksheets[4].cell(i, 2).value = '1,1,1-Trichloroethane (1,1,1-TCA)'
            both += 1
        if both == 9:
            break


    ## Non-Volatile SOCs

    both = 0
    for i in range(2, wb.worksheets[5].max_row + 1):
        if wb.worksheets[5].cell(i, 2).value == '2,4-D':
            wb.worksheets[5].cell(i, 2).value = '2,4-Dichlorophenoxyacetic acid (2,4-D)'
            both += 1
        elif wb.worksheets[5].cell(i, 2).value == 'Dalapon (qualitative)':
            wb.worksheets[5].cell(i, 2).value = 'Dalapon'
            both += 1
        elif wb.worksheets[5].cell(i, 2).value == 'PCP':
            wb.worksheets[5].cell(i, 2).value = 'Pentachlorophenol (PCP)'
            both += 1
        if both == 3:
            break


    ## Disinfection Byproducts

    ## Constituents with Notfi Levels

    # Changes the names respectively
    both = 0
    for i in range(2, wb.worksheets[7].max_row + 1):
        if wb.worksheets[7].cell(i, 2).value == 'o-Chlorotoluene':
            wb.worksheets[7].cell(i, 2).value = '2-Chlorotoluene'
            both += 1
        elif wb.worksheets[7].cell(i, 2).value == 'p-Chlorotoluene':
            wb.worksheets[7].cell(i, 2).value = '4-Chlorotoluene'
            both += 1
        elif wb.worksheets[7].cell(i, 2).value == 'PFOA':
            wb.worksheets[7].cell(i, 2).value = 'Perfluorooctanoic acid (PFOA)'
            both += 1
        elif wb.worksheets[7].cell(i, 2).value == 'PFOS':
            wb.worksheets[7].cell(i, 2).value = 'Perfluorooctanesulfonic acid (PFOS)'
            both += 1
        elif wb.worksheets[7].cell(i, 2).value == 'PFHxS':
            wb.worksheets[7].cell(i, 2).value = 'Perfluorohexane Sulfonic Acid (PFHxS)'
            both += 1
        elif wb.worksheets[7].cell(i, 2).value == 'PFBS':
            wb.worksheets[7].cell(i, 2).value = 'Perfluorobutane sulfonic acid (PFBS)'
            both += 1
        if both == 6:
            break

    wb.save(join(output_dir, ofile))











    ##quality control
    # q = dfs['MFSG'].loc[dfs['MFSG']['Category'] == 'Acid Extractables']


    ## write to file
    # for k in dfs.keys():
        # dfs[k].to_csv(path.join(data_dir, f'{k}_sorted_longform_v11.csv'), float_format="%.5f")
    # dfs['ARC WDR'].to_csv(path.join(data_dir, f'ARC WDR_sorted_longform_v11.csv'), float_format="%.6f")

    # for i in final_df['Site'].unique():
    #     print(f'{i} length:', len(final_df.loc[final_df['Site'] == i]))


### GRAVEYARD ###


# for site in x.columns.get_level_values(0).unique():
#     # print(site)
#     # temp_df = DataFrame(x[site].loc[x[site]['Action'] == 25]['Action'])          ##can drop
#     temp_df = DataFrame(x[site]['Action'].value_counts(), dtype = int64)
#     temp_df.columns = [site]
#     # temp_df.drop(temp_df.loc[0], inplace = True)
#     # temp2 = DataFrame(temp_df.count())
#     # temp2.columns = [site]
#     df = concat([df, temp_df], axis = 1)
#     # df2 = concat([df2, temp2], axis = 1).sum(axis = 1)
# df.drop(0, inplace = True)
#
# temp2 = []
# df2 = DataFrame()
# x2 = sd['ABP WDR']
# x2.drop(x2.iloc[:,0:2], axis = 1, inplace = True)
# for site in x2.columns.get_level_values(0).unique():
#     print(x2[site].iloc[:,0])
#     temp2.append([site, x2[site].iloc[:,0].count()])
#     # df2 = concat([df2, temp2], axis=1)
#     df2 = DataFrame(temp2)
# df2.columns = ['Site', 'Constituent Count']
# df2.to_excel(path.join(data_dir, 'Stat_Constituents_To_Review_Count.xlsx'))



# df.T.to_excel(path.join(data_dir, 'Cumulative_Actions.xlsx'))




##### XTRAS #####


    ## originally in merge_data() function, split to read sorting into beginning of script

    # sorting = read_csv(path.join(input_dir, 'class_constituent_p9.csv'), index_col = 1)
    # sorting['Constituents']= sorting['Constituents'].str.strip()
    # sorting.sort_values(by = 'Constituents', inplace = True)
    # dfs['ABP WDR'].index = dfs['ABP WDR'].iloc[:,0].str.split('(').str[0].str.strip()


    # dfs['ABP WDR'][0].fillna(dfs['ABP WDR']['index'], inplace=True)
    # dfs['ABP WDR'].sort_values(by = ['Category',0], inplace = True)
    # dfs['ABP WDR'].index = MultiIndex.from_frame(dfs['ABP WDR'].loc[:,['Category', 0]])
    # dfs['ABP WDR'].drop(['index', 'Category', 0], axis = 1, inplace=True)
